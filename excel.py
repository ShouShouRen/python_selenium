import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select

options = Options()
options.chrome_executable_path = "./chromedriver"
driver = webdriver.Chrome(options=options)

workbook = load_workbook('table.xlsx')
sheet = workbook.active

driver.get("http://localhost:5500/index.html")
input_elements = driver.find_elements("xpath", "//input[@type='text']")
select_element = driver.find_element("xpath", "//select[@name='select']")
yes_radio = driver.find_element("xpath", "//input[@id='yes']")
no_radio = driver.find_element("xpath", "//input[@id='no']")
for i, input_element in enumerate(input_elements):
    cell_value = sheet.cell(row=1, column=i+1).value
    input_element.send_keys(cell_value)
option_text = sheet['E1'].value
select = Select(select_element)
select.select_by_visible_text(option_text)
radio_value = sheet['F1'].value
if radio_value == "是":
    yes_radio.click()
else:
    no_radio.click()

time.sleep(5)
driver.quit()
workbook.close()
